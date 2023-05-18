import os
from aiogram.types import InputFile
from aiogram import types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.utils import executor
from aiogram import Bot, Dispatcher
import openpyxl
import io

# Токен
TOKEN = '6281162664:AAGTGqQ_wvDTUfq_FfOxJ6kMeYhbUOTuqWs'




# Подключение бота к телеграму
bot = Bot(token=TOKEN)
dp = Dispatcher(bot, storage=MemoryStorage())

print("Работаю!")


@dp.message_handler(commands=['start'])
async def bot_start(msg: types.Message):
    await bot.send_message(msg.from_user.id, f'Здравствуйте, {msg.from_user.full_name}! Я Бот для преобразования '
                                             f'excel-файлов. Отправь мне excel-файл')


@dp.message_handler(content_types=['document'])
async def doc_handler(message: types.Message):
    file_in_io = io.BytesIO()
    await message.reply(text='файл получен, начинаю поиск ошибок...')
    if message.content_type == 'document':
        destination = f"excel/{message.document.file_id}.xlsx"
        await message.document.download(destination_file=file_in_io, destination=destination)
    await message.reply(text='Обработка началась')

    # Обработка файла
    wb1 = openpyxl.load_workbook(f"excel/{message.document.file_id}.xlsx", data_only=True)
    worksheet1 = wb1.worksheets[0]
    wb2 = openpyxl.load_workbook("excel/Рассчет.xlsx", data_only=True)
    worksheet2 = wb2.worksheets[0]
    # Добавление информации в файл на сервере
    for row in worksheet1.iter_rows(values_only=True):
        if row[0] and row[0] != "№":
            worksheet2.append(row)

    wb2.save("excel/Рассчет.xlsx")  # Сохранение изменений в файле
    wb1.close()
    wb2.close()

    wb22 = openpyxl.load_workbook("excel/Рассчет.xlsx", data_only=True)
    # Создание нового excel файла
    wb3 = openpyxl.Workbook()
    wb3.create_sheet(wb22.sheetnames[1])
    wb3.create_sheet(wb22.sheetnames[2])

    try:
        if "Sheet" in wb3.sheetnames:
            wb3.remove_sheet(wb3.get_sheet_by_name("Sheet"))
    except Exception:
        pass

    worksheet22 = wb22.worksheets[1]
    worksheet3 = wb3.worksheets[0]
    # Добавление данных в новый файл

    for i in range(1, worksheet22.max_row + 1):
        for j in range(1, worksheet22.max_column + 1):
            c1 = worksheet22.cell(row=i, column=j)
            worksheet3.cell(row=i, column=j).value = c1.value

    worksheet22 = wb22.worksheets[2]
    worksheet4 = wb3.worksheets[1]
    for i in range(1, worksheet22.max_row + 1):
        for j in range(1, worksheet22.max_column + 1):
            c1 = worksheet22.cell(row=i, column=j)
            worksheet4.cell(row=i, column=j).value = c1.value
    wb3.close()
    wb22.close()
    wb3.save(f'excel/{message.document.file_id}_new.xlsx')  # Сохранение изменений в файле
    await message.answer_document(InputFile(f'excel/{message.document.file_id}_new.xlsx'))
    if os.path.isfile(f'excel/{message.document.file_id}_new.xlsx'):
        os.remove(f'excel/{message.document.file_id}_new.xlsx')
    if os.path.isfile(f"excel/{message.document.file_id}.xlsx"):
        os.remove(f"excel/{message.document.file_id}.xlsx")


if __name__ == '__main__':
    executor.start_polling(dp)
